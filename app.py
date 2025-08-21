import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import io
import openpyxl
from openpyxl.cell.cell import Cell
import os
from urllib.parse import urlparse
from pathlib import Path


def save_image(img_url, filename_prefix):
    if not img_url:
        return None
    try:
        downloads_dir = Path.home() / "Downloads"
        downloads_dir.mkdir(exist_ok=True)

        # Determine extension from URL, fallback to .jpg
        ext = os.path.splitext(urlparse(img_url).path)[1] or ".jpg"
        safe_prefix = "".join(c for c in filename_prefix if c.isalnum() or c in (' ', '_', '-')).strip()
        file_path = downloads_dir / f"{safe_prefix}{ext}"

        resp = requests.get(img_url, timeout=10)
        resp.raise_for_status()

        with open(file_path, "wb") as f:
            f.write(resp.content)

        return str(file_path)
    except Exception as e:
        print(f"Error saving {img_url}: {e}")
        return None


# --------------------
# Site-specific scrapers
# --------------------

def Fastline(url_list, progress_callback=None):
    equipment_list = []
    total = len(url_list)
    for i, link in enumerate(url_list):
        html = requests.get(link)
        x = BeautifulSoup(html.content , 'html.parser')
        equipment_dictionary = {}

        title_tag = x.find('title')
        title_text = title_tag.get_text(strip=True)[5:] if title_tag else ''
        equipment_dictionary["Title"] = title_text

        image_div = x.find('div', class_='item', attrs={'data-index': '0'})
        image_src = image_div.img['src'] if image_div and image_div.img else ''
        equipment_dictionary["Image"] = f'=IMAGE("{image_src}")' if image_src else ''
        equipment_dictionary["Local Image Path"] = save_image(image_src, f"{i+1}_{title_text[:50]}")

        def get_text(tag_label):
            tag = x.find('b', string=tag_label)
            return tag.next_sibling.strip() if tag and tag.next_sibling else ''

        year = get_text('Year:')
        make = get_text('Make:')
        model = get_text('Model:')
        hours = get_text('Hours:')
        miles = get_text('Mileage:')

        equipment_dictionary["Year/Make/Model"] = f"{year} {make} {model}".strip()
        if hours and miles:
            equipment_dictionary["Hours/Miles"] = f"{hours} HOURS, {miles} MILES"
        elif hours:
            equipment_dictionary["Hours/Miles"] = f"{hours} HOURS"
        elif miles:
            equipment_dictionary["Hours/Miles"] = f"{miles} MILES"
        else:
            equipment_dictionary["Hours/Miles"] = ''

        equipment_list.append(equipment_dictionary)
        if progress_callback:
            progress_callback((i+1)/total)
    return equipment_list



def Proxi_Bid(url_list, progress_callback=None):
    equipment_list = []
    total = len(url_list)
    for i, link in enumerate(url_list):
        html = requests.get(link)
        x = BeautifulSoup(html.content , 'html.parser')
        equipment_dictionary = {}

        title_tag = x.find('title')
        equipment_dictionary["Title"] = title_tag.get_text(strip=True).split('|')[0][7:] if title_tag else ''

        js_content = ''.join([s.string for s in x.find_all('script') if s.string])
        thumb_key = 'thumbnail: "'
        start = js_content.find(thumb_key)
        if start != -1:
            start += len(thumb_key)
            end = js_content.find('"', start)
            if end != -1:
                image_url = js_content[start:end]
                equipment_dictionary["Image"] = f'=IMAGE("{image_url}")'
            else:
                equipment_dictionary["Image"] = ''
        else:
            equipment_dictionary["Image"] = ''

        equipment_list.append(equipment_dictionary)

        if progress_callback:
            progress_callback((i+1)/total)
    return equipment_list


def Assiter(url_list, progress_callback=None):
    equipment_list = []
    total = len(url_list)
    for i, link in enumerate(url_list):
        html = requests.get(link)
        x = BeautifulSoup(html.content, 'html.parser')
        equipment_dictionary = {}

        title_str = x.find('title').get_text(strip=True)[22:] if x.find('title') else ''
        title_parts = title_str.split('  ')
        if len(title_parts) >= 3:
            equipment_dictionary["Year/Make/Model"] = f"{title_parts[0]} {title_parts[1]} {title_parts[2]}"
        else:
            equipment_dictionary["Year/Make/Model"] = ''

        description = x.find('meta', property='og:description')
        miles = ''
        if description:
            for line in str(description).split('\n'):
                if "Odo Reads" in line:
                    miles = line.split(':')[1].strip()
                    break
        equipment_dictionary["Hours/Miles"] = f'{miles} MILES' if miles else ''

        image_list = x.find_all('div', class_='image-gallery-image')
        image_link = ''
        for image in image_list:
            if 'play-button' not in str(image):
                image_link = str(image).split('"')[-2]
                break
        equipment_dictionary["Image"] = f'=IMAGE("{image_link}")' if image_link else ''

        equipment_list.append(equipment_dictionary)

        if progress_callback:
            progress_callback((i+1)/total)
    return equipment_list


def Kerr_Mowrey_Witcher(url_list, progress_callback=None):
    equipment_list = []
    total = len(url_list)
    for i, link in enumerate(url_list):
        html = requests.get(link)
        x = BeautifulSoup(html.content, 'html.parser')
        equipment_dictionary = {}

        title = x.find('meta', property="og:title")
        image = x.find('meta', property="og:image")

        equipment_dictionary["Title"] = title.get('content') if title else ''
        equipment_dictionary["Image"] = f'=IMAGE("{image.get("content")}")' if image else ''

        equipment_list.append(equipment_dictionary)

        if progress_callback:
            progress_callback((i+1)/total)
    return equipment_list


def Wausau(url_list, progress_callback=None):
    equipment_list = []
    total = len(url_list)
    for i, link in enumerate(url_list):
        html = requests.get(link)
        x = BeautifulSoup(html.content, 'html.parser')
        equipment_dictionary = {}

        desc = x.find('meta', property="og:description")
        title = desc.get('content').split(',')[0].strip() if desc else ''
        equipment_dictionary["Title"] = title

        image = x.find('meta', property="og:image")
        equipment_dictionary["Image"] = f'=IMAGE("{image.get("content")}")' if image else ''

        equipment_list.append(equipment_dictionary)

        if progress_callback:
            progress_callback((i+1)/total)
    return equipment_list













# --------------------
# Streamlit App
# --------------------

st.title("Equipment Scraper App")
st.markdown("Instructions")
st.markdown("""
            1. Upload a CSV file with one url per row (If using google sheets, download as CSV).
            2. Select the website from the dropdown
            3. Click on the "Begin" button to start the scraping process.
            4. Download the results as an Excel file.
            """)

uploaded_file = st.file_uploader("Upload a CSV file of URLs", type=["csv"])
website = st.selectbox("Select website type", ["Fastline", "Proxi_Bid", "Assiter", "Kerr", "Mowrey", "Witcher", "Wausau", "Quarrick", "Superior Energy"])

if uploaded_file:
    df_input = pd.read_csv(uploaded_file, header=None)
    url_list = df_input.iloc[:, 0].tolist()
    st.markdown("Preview of URLs:") #previewing for sanity check
    st.write(url_list[:3])

    if st.button("Begin"):
        st.subheader("Scraping Progress")
        st.caption(f"Processing {len(url_list)} URLs...")
        
        progress_bar = st.progress(0)
        
        def update_progress(progress):
            progress_bar.progress(progress)

        with st.spinner("Scraping in progress..."):
            if website == 'Fastline':
                equipment_list = Fastline(url_list, update_progress)
            elif website == "Proxi_Bid":
                equipment_list = Proxi_Bid(url_list, update_progress)
            elif website == "Assiter":
                equipment_list = Assiter(url_list, update_progress)
            elif website in ["Kerr", "Mowrey", "Witcher"]:
                equipment_list = Kerr_Mowrey_Witcher(url_list, update_progress)
            elif website == "Wausau":
                equipment_list = Wausau(url_list, update_progress)
            elif website == 'Quarrick':
                equipment_list = Proxi_Bid(url_list, update_progress)
            elif website == 'Superior Energy':
                equipment_list = Proxi_Bid(url_list, update_progress)

        df_output = pd.DataFrame(equipment_list)
        st.success("Scraping complete!")
        st.dataframe(df_output)

        template_path = "scraping_template1.xltm"  # Your macro-enabled Excel template

        # Load the template workbook
        wb = openpyxl.load_workbook(template_path, keep_vba=True)

        # Select the sheet where you want to write the data
        ws = wb["Sheet1"]  # Adjust if your sheet has a different name

        # Optional: Clear existing rows starting from row 2
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Write DataFrame to sheet starting at row 2
        for i, row in enumerate(df_output.values, start=2):
            for j, value in enumerate(row, start=1):
                if isinstance(value, str) and value.startswith('=IMAGE('):
                    ws.cell(row=i, column=j).value  = value
                else:
                    ws.cell(row=i, column=j, value = value)

        # Save to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Download button
        st.download_button(
            label="Download Excel with Macros",
            data=excel_buffer,
            file_name="scraped_equipment.xltm",
            mime="application/vnd.ms-excel.sheet.macroEnabled.12"
        )

    

        
        
        
        
        
        
        
        
        
        
        ''' excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
            df_output.to_excel(writer, index=False, sheet_name="Scraped Data")

        excel_buffer.seek(0)

        st.download_button(
            label="Download Excel",
            data=excel_buffer,
            file_name="scraped_equipment.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )'''
